#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import (
    compact_text,
    float_or_none,
    load_config,
    quartile_from_category,
    read_json,
    resolve_path,
    resolve_path_within,
    skill_root,
    workspace_output_path,
    write_json,
    write_text,
)


REQUIRED_HEADERS = {
    "journal_name",
    "issn",
    "eissn",
    "category",
    "citations",
    "jif",
    "jci",
    "percentageOAGold",
}


def load_catalog(catalog_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(catalog_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(item).strip() if item is not None else "" for item in next(rows)]
    missing = REQUIRED_HEADERS.difference(headers)
    if missing:
        raise ValueError(f"catalog 缺少列: {sorted(missing)}")

    catalog: list[dict[str, Any]] = []
    for row_values in rows:
        row = dict(zip(headers, row_values))
        catalog.append(
            {
                "journal_name": str(row.get("journal_name") or "").strip(),
                "issn": str(row.get("issn") or "").strip(),
                "eissn": str(row.get("eissn") or "").strip(),
                "category": str(row.get("category") or "").strip(),
                "citations": row.get("citations"),
                "jif": float_or_none(row.get("jif")),
                "jci": float_or_none(row.get("jci")),
                "percentage_oa_gold": float_or_none(row.get("percentageOAGold")),
            }
        )
    return catalog


def quartile_rank(quartile: str) -> int:
    return {"Q1": 0, "Q2": 1, "Q3": 2, "Q4": 3}.get(quartile, 4)


def build_selection_notes(row: dict[str, Any], profile: dict[str, Any], config: dict[str, Any]) -> list[str]:
    quartile = quartile_from_category(row["category"])
    notes = [f"JIF {row['jif']:.1f} 达到阈值", f"分区信号：{quartile}"]
    if profile.get("prefer_open_access") and (row["percentage_oa_gold"] or 0.0) >= float(
        config["screening"]["open_access_gold_threshold"]
    ):
        notes.append(f"OA Gold 占比 {row['percentage_oa_gold']:.1f}%")
    if row.get("citations"):
        notes.append(f"引用量 {row['citations']}")
    return notes


def summarize_candidate(row: dict[str, Any], profile: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    quartile = quartile_from_category(row["category"])
    return {
        **row,
        "quartile": quartile,
        "selection_notes": build_selection_notes(row, profile, config),
        "manual_exception": False,
    }


def sort_candidates(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        candidates,
        key=lambda item: (
            0 if item.get("manual_exception") else 1,
            quartile_rank(str(item.get("quartile") or "unknown")),
            -(item.get("jif") or 0.0),
            -(float_or_none(item.get("citations")) or 0.0),
            str(item.get("journal_name") or "").lower(),
        ),
    )


def apply_manual_exceptions(
    *,
    catalog: list[dict[str, Any]],
    shortlisted: list[dict[str, Any]],
    exceptions: list[dict[str, Any]] | None,
    profile: dict[str, Any],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    if not exceptions:
        return shortlisted
    journal_index = {item["journal_name"].lower(): item for item in catalog}
    deduped: dict[str, dict[str, Any]] = {
        item["journal_name"].lower(): item for item in shortlisted
    }
    for exception in exceptions:
        journal_name = str(exception.get("journal_name") or "").strip()
        if not journal_name:
            continue
        row = journal_index.get(journal_name.lower())
        if not row:
            continue
        candidate = {
            **summarize_candidate(row, profile, config),
            "manual_exception": True,
            "manual_exception_reason": str(exception.get("reason") or "").strip(),
        }
        deduped[journal_name.lower()] = candidate
    return sort_candidates(list(deduped.values()))


def build_shortlist(
    *,
    catalog: list[dict[str, Any]],
    profile: dict[str, Any],
    config: dict[str, Any],
    limit: int,
) -> list[dict[str, Any]]:
    if limit <= 0:
        raise ValueError(f"limit 必须是正整数，当前值为 {limit}")
    min_jif = max(float(config["screening"]["min_impact_factor"]), float(profile.get("target_if_min", 0.0) or 0.0))
    excluded = {
        str(name).strip().lower()
        for name in profile.get("excluded_journals", [])
        if str(name).strip()
    }
    candidates: list[dict[str, Any]] = []
    for row in catalog:
        if not row["journal_name"]:
            continue
        if row["journal_name"].lower() in excluded:
            continue
        if row["jif"] is None or row["jif"] < min_jif:
            continue
        candidates.append(summarize_candidate(row, profile, config))
    return sort_candidates(candidates)[:limit]


def render_markdown(payload: dict[str, Any]) -> str:
    lines = [
        "# Set1 候选池",
        "",
        f"- 生成时间：{payload['generated_at']}",
        f"- 候选数量：{payload['candidate_count']}",
        f"- 最低影响因子阈值：{payload['min_impact_factor']}",
        "- 本步骤只做最小硬过滤，不代表已完成语义匹配排序。",
        "",
        "| 排名 | 期刊 | 类别 | JIF | 分区 | 入池理由 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for index, item in enumerate(payload["candidates"], start=1):
        clue_text = "；".join(item.get("selection_notes") or []) or "待 AI 进一步判断"
        if item.get("manual_exception") and item.get("manual_exception_reason"):
            clue_text = f"{clue_text}；人工例外：{item['manual_exception_reason']}"
        lines.append(
            "| {idx} | {name} | {category} | {jif:.1f} | {quartile} | {clues} |".format(
                idx=index,
                name=item["journal_name"],
                category=item["category"].replace("|", "/"),
                jif=item["jif"] or 0.0,
                quartile=item.get("quartile", "unknown"),
                clues=clue_text.replace("|", "/"),
            )
        )
    lines.extend(
        [
            "",
            "## 下一步",
            "",
            "- 先通读 manuscript 画像与当前候选池，自主决定优先核验顺序。",
            "- 对真正值得继续看的候选联网核验：官方 scope、官网链接、业内认可度、预警/垃圾期刊风险。",
            "- 只把 scope 真匹配且质量信号干净的期刊写入 `analysis/set2_scope_review.json`。",
            "- 不要把当前 Set1 候选池直接当最终推荐清单。",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="基于 2023IF.xlsx 生成 Set1 候选池")
    parser.add_argument("--profile", required=True, help="manuscript_profile.json 路径")
    parser.add_argument("--workspace", required=True, help="隐藏工作区 run 根目录")
    parser.add_argument("--catalog", default="", help="覆盖默认内置 2023IF.xlsx")
    parser.add_argument("--limit", type=int, default=0, help="Set1 数量上限")
    parser.add_argument("--manual-include-json", default="", help="低 IF 例外补录 JSON")
    args = parser.parse_args()

    config = load_config()
    workspace = resolve_path(args.workspace)
    profile = read_json(resolve_path_within(args.profile, parent=workspace, label="profile"))
    catalog_path = (
        resolve_path(args.catalog)
        if args.catalog
        else (skill_root() / config["assets"]["catalog_xlsx"]).resolve()
    )
    catalog = load_catalog(catalog_path)
    limit = args.limit or int(config["screening"]["default_set1_limit"])
    shortlisted = build_shortlist(
        catalog=catalog,
        profile=profile,
        config=config,
        limit=limit,
    )
    if args.manual_include_json:
        shortlisted = apply_manual_exceptions(
            catalog=catalog,
            shortlisted=shortlisted,
            exceptions=read_json(resolve_path(args.manual_include_json)),
            profile=profile,
            config=config,
        )[:limit]

    payload = {
        "generated_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "catalog_path": str(catalog_path),
        "profile_title": profile.get("title", ""),
        "candidate_count": len(shortlisted),
        "min_impact_factor": max(
            float(config["screening"]["min_impact_factor"]),
            float(profile.get("target_if_min", 0.0) or 0.0),
        ),
        "candidates": shortlisted,
        "notes": compact_text(profile.get("notes", "")),
        "selection_strategy": "minimal-hard-filtering",
    }
    json_path = workspace_output_path(workspace, config["reports"]["set1_json"])
    markdown_path = workspace_output_path(workspace, config["reports"]["set1_markdown"])
    write_json(json_path, payload)
    write_text(markdown_path, render_markdown(payload))
    print(f"set1_json={json_path}")
    print(f"set1_markdown={markdown_path}")


if __name__ == "__main__":
    main()
